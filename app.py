from flask import Flask, render_template, request, send_file, flash, redirect, url_for, jsonify
import cv2
import numpy as np
import pytesseract
import pandas as pd
from tensorflow.keras.models import load_model
import os
from werkzeug.utils import secure_filename
from datetime import datetime
import re
import traceback

app = Flask(__name__)
app.secret_key = 'sv3-attendance-secret-key-2025'
UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'bmp', 'tiff'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Create folders
for folder in [UPLOAD_FOLDER, OUTPUT_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# Load model (optional)
try:
    if os.path.exists("model.h5"):
        model = load_model("model.h5")
        print("✓ Model loaded successfully")
    else:
        model = None
        print("⚠️ No model found, using rule-based detection")
except Exception as e:
    print(f"ERROR loading model: {e}")
    model = None

def find_and_straighten_sheet(image):
    """Detect and straighten the attendance sheet"""
    try:
        orig_h, orig_w = image.shape[:2]
        ratio = orig_h / 1000.0
        img_resized = cv2.resize(image, (int(orig_w / ratio), 1000))
        gray = cv2.cvtColor(img_resized, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blurred, 50, 150)
        
        kernel = np.ones((5,5), np.uint8)
        dilated = cv2.dilate(edges, kernel, iterations=2)
        
        contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        contours = sorted(contours, key=cv2.contourArea, reverse=True)[:5]
        
        screen_cnt = None
        for c in contours:
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.02 * peri, True)
            if len(approx) == 4:
                area = cv2.contourArea(approx)
                img_area = img_resized.shape[0] * img_resized.shape[1]
                if area > img_area * 0.3:
                    screen_cnt = approx
                    break
        
        if screen_cnt is None:
            print("⚠️ No suitable 4-point contour found, returning original image.")
            return image
        
        pts = screen_cnt.reshape(4, 2) * ratio
        rect = np.zeros((4, 2), dtype="float32")
        s = pts.sum(axis=1)
        rect[0] = pts[np.argmin(s)]
        rect[2] = pts[np.argmax(s)]
        diff = np.diff(pts, axis=1)
        rect[1] = pts[np.argmin(diff)]
        rect[3] = pts[np.argmax(diff)]
        
        (tl, tr, br, bl) = rect
        widthA = np.sqrt(((br[0] - bl[0]) ** 2) + ((br[1] - bl[1]) ** 2))
        widthB = np.sqrt(((tr[0] - tl[0]) ** 2) + ((tr[1] - tl[1]) ** 2))
        maxWidth = max(int(widthA), int(widthB))
        heightA = np.sqrt(((tr[0] - br[0]) ** 2) + ((tr[1] - br[1]) ** 2))
        heightB = np.sqrt(((tl[0] - bl[0]) ** 2) + ((tl[1] - bl[1]) ** 2))
        maxHeight = max(int(heightA), int(heightB))
        
        dst = np.array([[0, 0], [maxWidth - 1, 0], [maxWidth - 1, maxHeight - 1], [0, maxHeight - 1]], dtype="float32")
        M = cv2.getPerspectiveTransform(rect, dst)
        warped = cv2.warpPerspective(image, M, (maxWidth, maxHeight))
        print("✓ Applied perspective correction successfully.")
        return warped
    except Exception as e:
        print(f"Perspective correction failed: {e}")
        return image

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def detect_table_structure(img):
    """Detect grid lines to find table structure"""
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                   cv2.THRESH_BINARY_INV, 15, 2)
    
    # Detect horizontal lines
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (50, 1))
    horizontal_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel)
    
    # Detect vertical lines
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 50))
    vertical_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel)
    
    return horizontal_lines, vertical_lines

def find_cell_grid(img):
    """Find individual cells in the table"""
    h_lines, v_lines = detect_table_structure(img)
    
    # Find horizontal line positions
    h_contours, _ = cv2.findContours(h_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    h_positions = sorted(set([cv2.boundingRect(c)[1] for c in h_contours]))
    
    # Find vertical line positions  
    v_contours, _ = cv2.findContours(v_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    v_positions = sorted(set([cv2.boundingRect(c)[0] for c in v_contours]))
    
    # Remove duplicates (lines close together)
    h_positions = [h_positions[0]] + [h for i, h in enumerate(h_positions[1:], 1) 
                                       if h - h_positions[i-1] > 30]
    v_positions = [v_positions[0]] + [v for i, v in enumerate(v_positions[1:], 1) 
                                       if v - v_positions[i-1] > 30]
    
    print(f"✓ Detected grid: {len(h_positions)} rows x {len(v_positions)} columns")
    return h_positions, v_positions

def clean_text(text):
    """Clean OCR text output"""
    text = text.replace('\n', ' ')
    text = ' '.join(text.split())
    return text.strip()

def extract_text_from_region(img, x, y, w, h, lang='eng'):
    """Extract text from a specific region with enhanced preprocessing"""
    try:
        region = img[y:y+h, x:x+w]
        
        if region.size == 0:
            return ""
        
        if len(region.shape) == 3:
            gray = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
        else:
            gray = region
        
        # Resize if too small for better OCR
        if gray.shape[0] < 40:
            scale = 40 / gray.shape[0]
            gray = cv2.resize(gray, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
        
        # Multiple preprocessing approaches
        # Approach 1: CLAHE + Denoise
        clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
        enhanced1 = clahe.apply(gray)
        denoised1 = cv2.fastNlMeansDenoising(enhanced1, None, 10, 7, 21)
        _, binary1 = cv2.threshold(denoised1, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        
        # Approach 2: Simple adaptive threshold
        binary2 = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                        cv2.THRESH_BINARY, 11, 2)
        
        # Try OCR on both and pick better result
        config1 = f'--psm 6 -l {lang} --oem 3'
        config2 = f'--psm 7 -l {lang} --oem 3'
        
        text1 = pytesseract.image_to_string(binary1, config=config1)
        text2 = pytesseract.image_to_string(binary2, config=config2)
        
        # Pick the longer, more meaningful result
        text1_clean = clean_text(text1)
        text2_clean = clean_text(text2)
        
        if len(text1_clean) > len(text2_clean):
            return text1_clean
        else:
            return text2_clean if text2_clean else text1_clean
        
    except Exception as e:
        print(f"OCR error in region: {e}")
        return ""

def detect_checkmark_or_cross(cell_img):
    """
    Improved detection for checkmarks (✓) and circles (Ø) 
    Returns 'P', 'A', or ''
    """
    if cell_img.size == 0:
        return ""
    
    if len(cell_img.shape) == 3:
        gray = cv2.cvtColor(cell_img, cv2.COLOR_BGR2GRAY)
    else:
        gray = cell_img
    
    # Multiple thresholding attempts for robustness
    _, binary1 = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    _, binary2 = cv2.threshold(gray, 127, 255, cv2.THRESH_BINARY_INV)
    
    # Use the one with better contrast
    binary = binary1 if cv2.countNonZero(binary1) > cv2.countNonZero(binary2) * 0.5 else binary2
    
    # Calculate ink ratio
    ink_pixels = cv2.countNonZero(binary)
    total_pixels = binary.shape[0] * binary.shape[1]
    
    if total_pixels == 0:
        return ""
    
    ink_ratio = ink_pixels / total_pixels
    
    # Empty if very little ink
    if ink_ratio < 0.04:
        return ""
    
    # Find contours
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if not contours:
        return ""
    
    # Get largest contour
    largest_contour = max(contours, key=cv2.contourArea)
    area = cv2.contourArea(largest_contour)
    
    if area < 30:  # Lowered threshold
        return ""
    
    perimeter = cv2.arcLength(largest_contour, True)
    
    if perimeter == 0:
        return ""
    
    # Calculate shape features
    circularity = 4 * np.pi * area / (perimeter * perimeter)
    
    # Bounding box aspect ratio
    x, y, w, h = cv2.boundingRect(largest_contour)
    aspect_ratio = w / float(h) if h > 0 else 0
    
    # Convex hull solidity
    hull = cv2.convexHull(largest_contour)
    hull_area = cv2.contourArea(hull)
    solidity = area / hull_area if hull_area > 0 else 0
    
    # Enhanced decision logic
    # Ø (circle) characteristics: high circularity, aspect ~1.0
    if circularity > 0.6 and 0.7 < aspect_ratio < 1.3:
        return "A"
    
    # Circle with slash through it (also Ø variant)
    if circularity > 0.5 and ink_ratio > 0.15:
        return "A"
    
    # ✓ (checkmark) characteristics: low solidity, elongated
    if solidity < 0.75 and aspect_ratio < 0.9:
        return "P"
    
    # If has significant ink but not circular, likely checkmark
    if ink_ratio > 0.06 and circularity < 0.5:
        return "P"
    
    # Default to present if there's ink
    return "P" if ink_ratio > 0.05 else ""

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        if 'file' not in request.files or not request.files['file'].filename:
            flash('No file selected', 'error')
            return redirect(url_for('index'))
        
        file = request.files['file']
        if not allowed_file(file.filename):
            flash('Invalid file type. Please upload an image file.', 'error')
            return redirect(url_for('index'))
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{timestamp}_{secure_filename(file.filename)}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            img = cv2.imread(filepath)
            if img is None:
                flash('Could not read image file.', 'error')
                return redirect(url_for('index'))
            
            print(f"✓ Image loaded: {img.shape}")
            
            # Straighten image
            img = find_and_straighten_sheet(img)
            
            # Detect table grid
            h_positions, v_positions = find_cell_grid(img)
            
            if len(h_positions) < 3 or len(v_positions) < 5:
                flash('Could not detect table structure. Please ensure the image is clear.', 'error')
                return redirect(url_for('index'))
            
            # STEP 1: Extract dates from row 0
            date_headers = []
            col_idx = 4  # Start after No, Khmer Name, English Name, Sex
            
            while col_idx < len(v_positions) - 1:
                x1 = v_positions[col_idx]
                x2 = v_positions[min(col_idx + 3, len(v_positions) - 1)]
                y1, y2 = h_positions[0], h_positions[1]
                
                date_text = extract_text_from_region(img, x1, y1, x2-x1, y2-y1, 'eng')
                
                date_match = re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', date_text)
                if date_match:
                    actual_date = date_match.group(0)
                else:
                    actual_date = f"Date{len(date_headers)+1}"
                
                date_headers.append({
                    'date': actual_date,
                    'start_col': col_idx,
                    'num_subjects': 0,
                    'subjects': []
                })
                
                col_idx += 3
            
            print(f"✓ Found {len(date_headers)} dates: {[d['date'] for d in date_headers]}")
            
            # STEP 2: Extract subject names from row 1
            for date_info in date_headers:
                start_col = date_info['start_col']
                
                for subj_offset in range(3):
                    subj_col = start_col + subj_offset
                    
                    if subj_col >= len(v_positions) - 1:
                        break
                    
                    x1, x2 = v_positions[subj_col], v_positions[subj_col + 1]
                    y1, y2 = h_positions[1], h_positions[2]
                    
                    subject_text = extract_text_from_region(img, x1, y1, x2-x1, y2-y1, 'eng')
                    
                    if subject_text and len(subject_text) > 0:
                        subject_text = subject_text.replace('|', 'I').replace('1', 'I')
                        date_info['subjects'].append(subject_text)
                        date_info['num_subjects'] += 1
                    else:
                        date_info['subjects'].append(f"S{subj_offset+1}")
                        date_info['num_subjects'] += 1
            
            for date_info in date_headers:
                print(f"  {date_info['date']}: {date_info['subjects']}")
            
            # STEP 3: Build column structure - exactly matching the sheet
            # First row: Dates (spanning 3 columns each)
            # Second row: Subjects under each date
            columns_level1 = ['', '']  # For No and Name
            columns_level2 = ['No', 'Name']
            column_mapping = []
            
            for date_info in date_headers:
                # Add date header (will span 3 columns in Excel)
                columns_level1.extend([date_info['date']] * len(date_info['subjects']))
                
                # Add subjects under this date
                for subject in date_info['subjects']:
                    columns_level2.append(subject)
                    column_mapping.append(date_info['start_col'] + len(column_mapping) - 2)
            
            # STEP 4: Extract student data (start from row 2)
            records = []
            data_start_row = 2
            
            for row_idx in range(data_start_row, min(len(h_positions) - 1, data_start_row + 30)):
                y1, y2 = h_positions[row_idx], h_positions[row_idx + 1]
                
                if y2 - y1 < 20:
                    continue
                
                # Extract No (column 0)
                no_text = extract_text_from_region(
                    img, v_positions[0], y1, 
                    v_positions[1] - v_positions[0], y2 - y1, 'eng'
                )
                
                # Extract English Name (column 2 - skip Khmer column)
                eng_name = extract_text_from_region(
                    img, v_positions[2], y1,
                    v_positions[3] - v_positions[2], y2 - y1, 'eng'
                )
                
                if not eng_name or len(eng_name) < 2:
                    continue
                
                record = [no_text, eng_name]
                
                # Extract attendance marks
                for date_info in date_headers:
                    for subj_offset in range(len(date_info['subjects'])):
                        grid_col = date_info['start_col'] + subj_offset
                        
                        if grid_col >= len(v_positions) - 1:
                            record.append("")
                            continue
                        
                        x1, x2 = v_positions[grid_col], v_positions[grid_col + 1]
                        
                        padding = 5
                        cell_img = img[
                            max(0, y1+padding):min(img.shape[0], y2-padding),
                            max(0, x1+padding):min(img.shape[1], x2-padding)
                        ]
                        
                        # Returns 'P', 'A', or ''
                        status = detect_checkmark_or_cross(cell_img)
                        record.append(status)
                
                records.append(record)
                print(f"✓ Processed: {no_text} - {eng_name[:30]}")
            
            if not records:
                flash('Could not extract student data. Please check image quality.', 'error')
                return redirect(url_for('index'))
            
            # Create simple DataFrame first
            df = pd.DataFrame(records, columns=columns_level2)
            
            print(f"✓ Created DataFrame with {len(df)} students")
            
            # Save to Excel
            excel_filename = f"attendance_{timestamp}.xlsx"
            excel_path = os.path.join(app.config['OUTPUT_FOLDER'], excel_filename)
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Attendance', index=False, startrow=1, header=False)
                worksheet = writer.sheets['Attendance']
                
                # Manually create multi-level header
                from openpyxl.styles import Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                # Write first header row (dates) - merged cells
                col_idx = 1
                
                # Merge No and Name cells vertically (rows 1-2)
                worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
                cell = worksheet.cell(1, 1, 'No')
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
                cell = worksheet.cell(1, 2, 'Name')
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                col_idx = 3
                
                # Write dates (spanning multiple columns)
                for date_info in date_headers:
                    num_subjects = len(date_info['subjects'])
                    if num_subjects > 1:
                        worksheet.merge_cells(start_row=1, start_column=col_idx, 
                                            end_row=1, end_column=col_idx+num_subjects-1)
                    
                    cell = worksheet.cell(1, col_idx, date_info['date'])
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    col_idx += num_subjects
                
                # Write second header row (subjects)
                col_idx = 3
                
                for date_info in date_headers:
                    for subject in date_info['subjects']:
                        cell = worksheet.cell(2, col_idx, subject)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        col_idx += 1
                
                # Set column widths
                worksheet.column_dimensions['A'].width = 8
                worksheet.column_dimensions['B'].width = 25
                for i in range(3, len(df.columns) + 1):
                    worksheet.column_dimensions[get_column_letter(i)].width = 12
                
                # Color code attendance cells
                from openpyxl.styles import PatternFill
                
                green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Data starts at row 3 (after 2 header rows)
                for row in worksheet.iter_rows(min_row=3, max_row=len(df)+2, 
                                               min_col=3, max_col=len(df.columns)):
                    for cell in row:
                        if cell.value == "P":
                            cell.fill = green_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        elif cell.value == "A":
                            cell.fill = red_fill
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            flash(f'✅ Successfully processed {len(df)} students!', 'success')
            
            return send_file(
                excel_path,
                as_attachment=True,
                download_name=excel_filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        
        except Exception as e:
            flash(f'An error occurred: {str(e)}', 'error')
            traceback.print_exc()
            return redirect(url_for('index'))
    
    return render_template("index.html", model_loaded=(model is not None))

@app.route("/health")
def health():
    return jsonify({
        "status": "ok",
        "model_loaded": model is not None,
        "timestamp": datetime.now().isoformat()
    })

@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404

@app.errorhandler(413)
def too_large(e):
    return "File is too large. Maximum size is 16MB.", 413

if __name__ == "__main__":
    print("\n" + "="*60)
    print("📋 SV3 Attendance Recognition System - SIMPLIFIED")
    print("="*60)
    print("Output Format:")
    print("  - No | Name | Date1_Subject1 | Date1_Subject2 | ...")
    print("  - ✓ = 'P' (Present)")
    print("  - Ø/A = 'A' (Absent)")
    print("  - Empty = '' (blank)")
    print("="*60 + "\n")
    app.run(debug=True, host='0.0.0.0', port=5000)